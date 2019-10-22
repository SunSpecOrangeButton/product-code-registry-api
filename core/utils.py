from openpyxl import load_workbook
from .serializers import *
import json


def excelToDict(excel_file):

    return_obj = {
        'module': None,
        'inverter': None,
        'battery': None,
    }
    wb = load_workbook(excel_file, read_only=True)
    print(wb.sheetnames)
    for sheet in wb.sheetnames:
        if sheet == "Modules":
            moduleProducts = []
            moduleSheet = wb[sheet]
            print("Module sheet max row:")
            print(moduleSheet.max_row)
            row_iterator = moduleSheet.values
            keys = next(row_iterator)
            # print(keys)
            for row in row_iterator:
                moduleProducts.append(dict(zip(keys, row)))
            #print(moduleProducts)
            module_serializer = ModuleSerializer(data=moduleProducts, many=True)

            if module_serializer.is_valid():
                module_serializer.save()
                return_obj['module'] = True
            else:
                print(module_serializer.errors)
        elif sheet == "Inverters":
            inverterProducts = []
            inverterSheet = wb[sheet]
            print("inverter sheet max row:")
            print(inverterSheet.max_row)
            row_iterator = inverterSheet.values
            keys = next(row_iterator)
            # print(keys)
            for row in row_iterator:
                inverterProducts.append(dict(zip(keys, row)))
            #print(inverterProducts)
            inverter_serializer = InverterSerializer(data=inverterProducts, many=True)
            if inverter_serializer.is_valid():
                inverter_serializer.save()
                return_obj['inverter'] = True
            else:
                print(inverter_serializer.errors)
        # elif sheet == "Batteries":
        #     pass
        else:
            print("Sheet is formatted incorrectly")
    return return_obj


def excelToDictGenericProduct(excel_file):
    successful_upload = False;
    wb = load_workbook(excel_file, read_only=True)
    productGenericSheet = wb.active
    productGenericList = []

    row_iterator = productGenericSheet.values
    keys = next(row_iterator)
    # print(keys)
    for row in row_iterator:
        productGenericList.append(dict(zip(keys, row)))
    #print(moduleProducts)
    product_serializer = ProductSerializer(data=productGenericList, many=True)

    if product_serializer.is_valid():
        product_serializer.save()
        successful_upload = True
    else:
        print(product_serializer.errors)

    return successful_upload


### Need to save serialized data of both Inverter and Module
# depending on whichever, or both is sent







